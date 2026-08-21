#!/usr/bin/env python3
"""Export every public PostgreSQL table to a formatted multi-sheet XLSX workbook."""

import argparse
import datetime as dt
import decimal
import json
import os
import re
import uuid

import psycopg2
from psycopg2 import sql
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="12355B")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
MAX_CELL_LENGTH = 32_000
MAX_DATA_ROWS_PER_SHEET = 1_048_575

SENSITIVE_PATTERNS = (
    "password",
    "passwd",
    "secret",
    "access_token",
    "refresh_token",
    "token_hash",
    "id_token",
    "api_key",
    "private_key",
    "client_secret",
    "authorization",
)
SENSITIVE_TABLE_COLUMNS = {
    ("settings", "value"),
    ("meta_kv", "value"),
    ("refresh_tokens", "token"),
}


def is_sensitive(table, column):
    lowered = column.lower()
    return (
        (table, column) in SENSITIVE_TABLE_COLUMNS
        or any(pattern in lowered for pattern in SENSITIVE_PATTERNS)
    )


def safe_excel_text(value):
    if isinstance(value, bytes):
        return f"[BINARY DATA: {len(value):,} bytes]"
    if isinstance(value, (dict, list, tuple)):
        value = json.dumps(value, ensure_ascii=False, default=str)
    elif isinstance(value, (dt.datetime, dt.date, dt.time)):
        value = value.isoformat()
    elif isinstance(value, (decimal.Decimal, uuid.UUID)):
        value = str(value)
    elif value is not None and not isinstance(value, (str, int, float, bool)):
        value = str(value)

    if isinstance(value, str):
        if len(value) > MAX_CELL_LENGTH:
            value = value[:MAX_CELL_LENGTH] + "\n[TRUNCATED FOR EXCEL]"
        if value.startswith(("=", "+", "-", "@")):
            value = "'" + value
    return value


def safe_sheet_name(table, used, part=None):
    base = re.sub(r"[\[\]:*?/\\]", "_", table)
    suffix = f"_{part}" if part else ""
    candidate = (base[: 31 - len(suffix)] + suffix) or "Sheet"
    counter = 2
    while candidate.lower() in used:
        extra = f"_{counter}"
        candidate = base[: 31 - len(extra)] + extra
        counter += 1
    used.add(candidate.lower())
    return candidate


def style_table_sheet(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 24
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sample_rows = min(ws.max_row, 201)
    for index in range(1, ws.max_column + 1):
        longest = len(str(ws.cell(1, index).value or ""))
        for row in range(2, sample_rows + 1):
            value = ws.cell(row, index).value
            if value is not None:
                lines = str(value).splitlines()[:3]
                longest = max(longest, max((len(line) for line in lines), default=0))
        ws.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 12), 45)


def style_index_sheet(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for column in range(1, ws.max_column + 1):
        values = [str(ws.cell(row, column).value or "") for row in range(1, min(ws.max_row, 300) + 1)]
        ws.column_dimensions[get_column_letter(column)].width = min(max(map(len, values)) + 2, 55)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("output")
    args = parser.parse_args()

    conn = psycopg2.connect(
        host=os.getenv("PGHOST", "127.0.0.1"),
        port=int(os.getenv("PGPORT", "5435")),
        user=os.getenv("PGUSER", "postgres"),
        password=os.getenv("PGPASSWORD", ""),
        dbname=os.getenv("PGDATABASE", "decoinks_db"),
    )
    conn.set_session(readonly=True, autocommit=False)

    with conn.cursor() as cursor:
        cursor.execute(
            """
            SELECT table_name
            FROM information_schema.tables
            WHERE table_schema='public' AND table_type='BASE TABLE'
            ORDER BY table_name
            """
        )
        tables = [row[0] for row in cursor.fetchall()]
        cursor.execute(
            """
            SELECT table_name,column_name,ordinal_position,data_type,udt_name,
                   is_nullable,column_default
            FROM information_schema.columns
            WHERE table_schema='public'
            ORDER BY table_name,ordinal_position
            """
        )
        dictionary = cursor.fetchall()

    columns_by_table = {}
    for row in dictionary:
        columns_by_table.setdefault(row[0], []).append(row[1])

    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()

    readme = wb.create_sheet("README")
    used_names.add("readme")
    readme.append(["Decoinks Database Export"])
    readme["A1"].fill = HEADER_FILL
    readme["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    readme.merge_cells("A1:D1")
    readme.append(["Generated (UTC)", dt.datetime.now(dt.timezone.utc).isoformat()])
    readme.append(["Database", os.getenv("PGDATABASE", "decoinks_db")])
    readme.append(["Schema", "public"])
    readme.append(["Tables exported", len(tables)])
    readme.append(["Format", "One worksheet per table; very large tables are split into numbered parts."])
    readme.append(["Security", "Password, token, API-key, and secret values are masked as [REDACTED]. Column names remain visible."])
    readme.append(["Safety", "Text beginning with =, +, -, or @ is escaped to prevent spreadsheet formula execution."])
    readme.append(["Binary values", "Represented by byte length instead of embedding raw binary data."])
    readme.append(["Data Dictionary", "Lists every exported column, type, nullability, default, and masking status."])
    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 100
    readme.sheet_view.showGridLines = False

    summary = wb.create_sheet("Table Summary")
    used_names.add("table summary")
    summary.append(["Table", "Worksheet", "Rows", "Columns", "Sensitive columns masked"])

    data_dictionary = wb.create_sheet("Data Dictionary")
    used_names.add("data dictionary")
    data_dictionary.append(
        ["Schema", "Table", "Column", "Position", "Data Type", "PostgreSQL Type", "Nullable", "Default", "Sensitive/Masked"]
    )
    for table, column, position, data_type, udt_name, nullable, default in dictionary:
        data_dictionary.append(
            ["public", table, column, position, data_type, udt_name, nullable, default, "Yes" if is_sensitive(table, column) else "No"]
        )

    for table in tables:
        columns = columns_by_table.get(table, [])
        with conn.cursor() as cursor:
            cursor.execute(sql.SQL("SELECT COUNT(*) FROM {}.{}").format(sql.Identifier("public"), sql.Identifier(table)))
            row_count = cursor.fetchone()[0]

        sheet_names = []
        part = 1
        row_offset = 0
        while row_offset < max(row_count, 1):
            sheet_name = safe_sheet_name(table, used_names, part if row_count > MAX_DATA_ROWS_PER_SHEET else None)
            sheet_names.append(sheet_name)
            ws = wb.create_sheet(sheet_name)
            ws.append(columns)

            if row_count:
                cursor_name = f"export_{re.sub('[^a-zA-Z0-9_]', '_', table)}_{part}"
                with conn.cursor(name=cursor_name) as cursor:
                    cursor.itersize = 2_000
                    cursor.execute(
                        sql.SQL("SELECT * FROM {}.{} ORDER BY ctid LIMIT %s OFFSET %s").format(
                            sql.Identifier("public"), sql.Identifier(table)
                        ),
                        (MAX_DATA_ROWS_PER_SHEET, row_offset),
                    )
                    for row in cursor:
                        output_row = []
                        for column, value in zip(columns, row):
                            output_row.append("[REDACTED]" if value is not None and is_sensitive(table, column) else safe_excel_text(value))
                        ws.append(output_row)

            style_table_sheet(ws)
            row_offset += MAX_DATA_ROWS_PER_SHEET
            part += 1
            if row_count == 0:
                break

        sensitive_count = sum(is_sensitive(table, column) for column in columns)
        summary.append([table, ", ".join(sheet_names), row_count, len(columns), sensitive_count])

    style_index_sheet(summary)
    style_index_sheet(data_dictionary)
    summary.auto_filter.ref = summary.dimensions
    data_dictionary.auto_filter.ref = data_dictionary.dimensions

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    wb.save(args.output)
    conn.rollback()
    conn.close()

    check = load_workbook(args.output, read_only=True, data_only=False)
    print(
        json.dumps(
            {
                "output": os.path.abspath(args.output),
                "tables": len(tables),
                "worksheets": len(check.sheetnames),
                "worksheet_names": check.sheetnames,
                "file_size": os.path.getsize(args.output),
            },
            indent=2,
        )
    )
    check.close()


if __name__ == "__main__":
    main()
